"""
Report generation module.
Produces a formatted multi-sheet Excel workbook for download.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import polars as pl

# ── Style constants ───────────────────────────────────────────────────────────

HEADER_BG   = "1F4E79"   # dark blue
HEADER_FG   = "FFFFFF"   # white text
ALT_ROW_BG  = "EBF3FB"   # light blue for alternating rows
FLAG_RED_BG = "FFCCCC"   # red highlight for gap/obsolete flags
FLAG_RED_FG = "CC0000"

MAX_COL_WIDTH = 50
MIN_COL_WIDTH = 12


# ── Internal helpers ──────────────────────────────────────────────────────────

def _header_style():
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    font = Font(color=HEADER_FG, bold=True, size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return fill, font, align


def _flag_cols(df: pl.DataFrame) -> set[int]:
    """Return 1-based column indices for gap_* / is_obsolete columns."""
    flagged = set()
    for i, col in enumerate(df.columns, 1):
        if col.startswith("gap_") or col in ("is_obsolete", "retired_still_tagged"):
            flagged.add(i)
    return flagged


def _write_dataframe(ws, df: pl.DataFrame) -> None:
    """Write a Polars DataFrame into an openpyxl worksheet with formatting."""
    h_fill, h_font, h_align = _header_style()
    flag_indices = _flag_cols(df)
    red_fill = PatternFill(start_color=FLAG_RED_BG, end_color=FLAG_RED_BG, fill_type="solid")
    alt_fill = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type="solid")

    # Header row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = h_align

    # Data rows
    for row_idx, row in enumerate(df.iter_rows(), 2):
        is_alt = row_idx % 2 == 0
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
            # Highlight flag columns where value is True
            if col_idx in flag_indices and str(value).lower() == "true":
                cell.fill = red_fill
                cell.font = Font(color=FLAG_RED_FG, bold=True)
            elif is_alt:
                cell.fill = alt_fill

    # Auto-fit column widths
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        # Sample up to 200 rows to estimate width
        sample_values = df[col_name].head(200).cast(pl.Utf8).to_list()
        max_len = max(
            (len(str(v)) for v in sample_values if v),
            default=len(col_name),
        )
        width = max(MIN_COL_WIDTH, min(max_len + 2, MAX_COL_WIDTH))
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"


# ── Public API ────────────────────────────────────────────────────────────────

EXCEL_MAX_ROWS = 1_048_575  # 1 row reserved for header


def _iter_chunks(df: pl.DataFrame, chunk_size: int):
    """Yield successive chunk_size slices of df."""
    for start in range(0, len(df), chunk_size):
        yield df.slice(start, chunk_size)


def generate_report(sheets: dict[str, pl.DataFrame]) -> bytes:
    """
    Build a formatted Excel workbook with one sheet per entry in `sheets`.
    DataFrames exceeding Excel's 1,048,576 row limit are automatically
    split into numbered continuation sheets (Part 1, Part 2, …).
    Returns raw bytes suitable for st.download_button.
    """
    wb = Workbook()
    wb.remove(wb.active)  # remove the default blank sheet

    for sheet_name, df in sheets.items():
        if len(df) <= EXCEL_MAX_ROWS:
            ws = wb.create_sheet(title=sheet_name[:31])
            _write_dataframe(ws, df)
        else:
            # Split into chunks and write as Part 1, Part 2, ...
            for part, chunk in enumerate(_iter_chunks(df, EXCEL_MAX_ROWS), start=1):
                base = sheet_name[:24]   # leave room for " (Pt X)"
                title = f"{base} (Pt {part})"
                ws = wb.create_sheet(title=title)
                _write_dataframe(ws, chunk)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
